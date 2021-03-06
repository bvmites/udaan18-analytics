import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.append('./src/')
import helper


def add_formatting(worksheet):
    # First column bold
    for rows in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1):
        for cell in rows:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal='center')
    # First row bold
    for rows in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=worksheet.max_column):
        for cell in rows:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal='center')
    return worksheet


def write_to_excel(df_list, event_list, start_string):
    """
    Helper function to write a list of Pandas dataframe to excel file
    :param df_list: list of Pandas Dataframe
    :return:
    """
    i = 0
    print(event_list)
    event_list = [list(map(lambda x: x.replace("/", "_"), event)) for event in event_list]

    for df in df_list:
        # Shift Index by 1
        df.index += 1
        j = 0
        max_list = []  # Lis1t that stores max length of strings in a column

        for col in df.columns.values:
            max_list.append(
                max(len(col), df[col].map(str).map(len).max())
            )

        if not os.path.exists('../total_part/' + start_string + '_' + str(event_list[i][1]) + '.xlsx'):
            writer = pd.ExcelWriter("../total_part/" + start_string + '_' + str(event_list[i][1]) + '.xlsx',
                                    engine='openpyxl')

            df.to_excel(writer, sheet_name=str(event_list[i][2]))

            workbook = writer.book
            worksheet = writer.sheets[str(event_list[i][2])]
        else:
            workbook = load_workbook("../total_part/" + start_string + '_' + str(event_list[i][1]) + '.xlsx')

            # Check if sheet exists. Create a new one otherwise
            if str(event_list[i][2]) in workbook.get_sheet_names():
                worksheet = workbook[str(event_list[i][2])]
                df.index += worksheet.max_row - 1
                dataframe_rows = list(dataframe_to_rows(df))
                for row_count in range(1, len(dataframe_rows)):
                    worksheet.append(dataframe_rows[row_count])
            else:
                worksheet = workbook.create_sheet(str(event_list[i][2]))
                dataframe_rows = list(dataframe_to_rows(df))
                for row_count in range(0, len(dataframe_rows)):
                    worksheet.append(dataframe_rows[row_count])

        worksheet = add_formatting(worksheet)
        cols = list(worksheet.columns)
        for j in range(1, len(cols)):
            worksheet.column_dimensions[cols[j][0].column].width = max_list[j - 1]
            j += 1

        workbook.save("../total_part/" + start_string + '_' + str(event_list[i][1]) + '.xlsx')
        workbook.close()
        i += 1


def main():
    # Choose the db
    db = helper.init_db()

    # Init Mapping
    mapping = helper.get_mapping('fields_mapping.json')

    # Empty list for storing pandas DataFrames later
    df_list = []

    # List of Columns
    columns = ['Event Name', 'Total Participation']

    # Query
    # data = db.Events.aggregate([{"$project": {"_id": 1, "name": 1, "participants": 1}},
    #                             {"$unwind": "$participants"}, {"$group": {"_id": {"_id": "$_id", "name": "$name"},
    #                                                                       "count": {"$sum": 1}}}])

    events_output = sorted(
        (db.events.aggregate([{"$project": {"_id": 1, mapping["name"]: 1, mapping["fee"]: 1,
                                            "department": 1, "eventType": 1}}])),
        key=lambda k: k['_id']
    )
    cur_date = 8
    query_output = db.participations.aggregate([{"$group": {"_id": {"eventId": "$eventId", "date": {
        "$dayOfMonth": "$" + mapping["pRegDate"]}},
                                                            "count": {"$sum": 1}}},
                                                {"$match": {"_id.date": cur_date}}
                                                # {"$project": {"_id": 1, "count": 1}}
                                                ])

    query_output = sorted(list(query_output), key=lambda x: x['_id']['eventId'])
    i = 0
    event_list = []
    for event in query_output:
        if events_output[i]['_id'] == event['_id']['eventId']:
            # Append name to event_list
            if events_output[i]['eventType'].lower() != 'technical':
                events_output[i]['department'] = 'Sheet1'
            event_list.append(
                [events_output[i][mapping['name']],
                 events_output[i]['eventType'],
                 events_output[i]['department']]
            )
            df_list.append(
                pd.DataFrame(data=[[events_output[i]['eventName'], event['count']]], columns=columns)
            )

        i += 1
    write_to_excel(df_list, event_list, 'total_part')
    # Write to excel sheet
    # i = 0
    # for df in df_list:
    #     df.to_excel('df' + str(i) + '.xlsx')
    #     i += 1


if __name__ == '__main__':
    main()

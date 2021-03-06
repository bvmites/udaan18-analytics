import os
import pprint
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.append('./src/')
import helper


def add_formatting(worksheet):
    # First column bold
    for rows in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1):
        for cell in rows:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal='center')
    # First row bold
    for rows in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=worksheet.max_column):
        for cell in rows:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal='center')
    return worksheet


def write_to_excel(df_list, event_list, start_string):
    """
    Helper function to write a list of Pandas dataframe to excel file
    :param df_list: list of Pandas Dataframe
    :return:
    """
    i = 0
    print(event_list)
    event_list = [list(map(lambda x: x.replace("/", "_"), event)) for event in event_list]

    for df in df_list:
        # Shift Index by 1
        df.index += 1
        j = 0
        max_list = []  # Lis1t that stores max length of strings in a column

        for col in df.columns.values:
            max_list.append(
                max(len(col), df[col].map(str).map(len).max())
            )

        if not os.path.exists('../part_year/' + start_string + '_' + str(event_list[i][1]) + '.xlsx'):
            writer = pd.ExcelWriter("../part_year/" + start_string + '_' + str(event_list[i][1]) + '.xlsx',
                                    engine='openpyxl')

            df.to_excel(writer, sheet_name=str(event_list[i][2]))

            workbook = writer.book
            worksheet = writer.sheets[str(event_list[i][2])]
        else:
            workbook = load_workbook("../part_year/" + start_string + '_' + str(event_list[i][1]) + '.xlsx')

            # Check if sheet exists. Create a new one otherwise
            if str(event_list[i][2]) in workbook.get_sheet_names():
                worksheet = workbook[str(event_list[i][2])]
                df.index += worksheet.max_row - 1
                dataframe_rows = list(dataframe_to_rows(df))
                for row_count in range(1, len(dataframe_rows)):
                    worksheet.append(dataframe_rows[row_count])
            else:
                worksheet = workbook.create_sheet(str(event_list[i][2]))
                dataframe_rows = list(dataframe_to_rows(df))
                for row_count in range(0, len(dataframe_rows)):
                    worksheet.append(dataframe_rows[row_count])

        worksheet = add_formatting(worksheet)
        cols = list(worksheet.columns)
        for j in range(1, len(cols)):
            worksheet.column_dimensions[cols[j][0].column].width = max_list[j - 1]
            j += 1

        workbook.save("../part_year/" + start_string + '_' + str(event_list[i][1]) + '.xlsx')
        workbook.close()
        i += 1


def main():
    # Choose the db
    db = helper.init_db()

    # Init Mapping
    mapping = helper.get_mapping('fields_mapping.json')

    # Empty list for storing pandas DataFrames later
    df_list = []

    # List of Columns
    columns = ['Event Name', 'Year 1', 'Year 2', 'Year 3', 'Year 4']

    # Mapped Query
    # query_output = db.Events.aggregate([{"$project": {mapping['participants']: 1, mapping["name"]: 1}},
    #                                     {"$unwind": "$" + mapping["participants"]},
    #                                     {"$group": {
    #                                         "_id": {"eventID": "$_id",
    #                                                 "year": "$" + mapping['participants'] + "." + mapping["pYear"],
    #                                                 "name": "$" + mapping["name"]},
    #                                         "count": {"$sum": 1}}},
    #                                     {"$group": {"_id": {"eventID": "$_id.eventID", "name": "$_id.name"},
    #                                                 "entries": {"$push": {"year": "$_id.year", "count": "$count"}}}}])

    events_output = sorted(
        (db.events.aggregate([{"$project": {"_id": 1, mapping["name"]: 1, mapping["fee"]: 1,
                                            "department": 1, "eventType": 1}}])),
        key=lambda k: k['_id']
    )
    date = 3
    query_output = db.participations.aggregate([{"$group": {
        "_id": {"eventID": "$eventId",
                "year": "$year",
                "date": {
                    "$dayOfMonth": "$" + mapping["pRegDate"]}
                },
                                            "count": {"$sum": 1}}},
        {"$match": {"_id.date": date}},
        {"$group": {"_id": {"eventID": "$_id.eventID"},
                    "count_list": {"$push": {"year": "$_id.year", "count": "$count"}}}}])
    query_output = sorted(list(query_output), key=lambda x: x['_id']['eventID'])
    pp = pprint.PrettyPrinter(indent=4)
    pp.pprint(query_output)

    # Loop through events
    i = 0
    event_list = []
    for event in query_output:
        while True:
            if events_output[i]['_id'] == event['_id']['eventID']:
                # Append name to event_list
                if events_output[i]['eventType'].lower() != 'technical':
                    events_output[i]['department'] = 'Sheet1'
                event_list.append(
                    [events_output[i][mapping['name']],
                     events_output[i]['eventType'],
                     events_output[i]['department']]
                )
                year_list = [0] * 4
                for entry in sorted(event['count_list'], key=lambda x: x['year']):
                    year_list[int(entry['year']) - 1] = entry['count']
                year_list.insert(0, events_output[i]['eventName'])
                df_list.append(
                    pd.DataFrame(data=[year_list], columns=columns)
                )
                i += 1
                break
            i += 1
    i = 0

    write_to_excel(df_list, event_list, 'part_year')


if __name__ == '__main__':
    main()
